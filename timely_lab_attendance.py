# -*- coding: utf-8 -*-
"""
Created on Sun Jul 27 11:20:10 2025

@author: Drew Ritcher
"""

import openpyxl


def select_active_tab(workbook, tab_name):
    workbook.active = workbook[tab_name]


def calculate_number_of_labs_attended(active_sheet, tab_name):
    number_of_labs_column = 0
    
    for row in range(2, active_sheet.max_row + 1):
        lab_participation = 0
        student_name = active_sheet["A" + str(row)].value
        
        if student_name is None:
            continue
        
        if active_sheet["B" + str(row)].value not in main_student_hash.keys():
            main_student_hash[active_sheet["B" + str(row)].value] = {}
            main_student_hash[active_sheet["B" + str(row)].value]["student"] = student_name
            main_student_hash[active_sheet["B" + str(row)].value]["sis_user_id"] = active_sheet["C" + str(row)].value
            main_student_hash[active_sheet["B" + str(row)].value]["sis_login_id"] = active_sheet["D" + str(row)].value
            main_student_hash[active_sheet["B" + str(row)].value]["root_account"] = active_sheet["E" + str(row)].value
            main_student_hash[active_sheet["B" + str(row)].value]["section"] = active_sheet["F" + str(row)].value
        
        main_student_hash[active_sheet["B" + str(row)].value][tab_name] = {}
        
        if tab_name not in active_sheet.cell(row=1, column=active_sheet.max_column).value:
            number_of_labs_column = active_sheet.max_column + 1
            active_sheet.cell(row=1, column=number_of_labs_column).value = active_sheet.title + " # of labs attended"
            
            active_sheet_headers.insert(number_of_labs_column, active_sheet.title + " # of labs attended")
        else:
            number_of_labs_column = active_sheet.max_column
            
        
        for column in range(7, number_of_labs_column):
            grade = active_sheet.cell(row=row, column=column).value
            main_student_hash[active_sheet["B" + str(row)].value][tab_name][active_sheet_headers[column]] = grade

            if grade is not None:
                if grade > 0:   
                    lab_participation += 1
        
        
        active_sheet.cell(row=row, column=number_of_labs_column).value = lab_participation
        main_student_hash[active_sheet["B" + str(row)].value][tab_name][active_sheet_headers[column]] = lab_participation
        
    return main_student_hash


def lab_attendance_deduction_after_key_lab(num_labs_attended, expected_num_labs_attended):
    if num_labs_attended < expected_num_labs_attended:
        return 4
    else:
        return 0


def lab_points_earned_by_deadline(num_labs_attended, expected_num_labs_attended):
    if num_labs_attended >= expected_num_labs_attended:
        return 4
    else:
        return 0


# ============================================================================
#  === Change the key lab dates and the excel file (with extension) here ====
# ============================================================================

key_lab_dates = ['2-15', '3-8', '4-5', '4-19', '4-26']
lab_excel_file = "combined_lab_exercises.xlsx"

# ============================================================================
# ============================================================================
# ============================================================================


from openpyxl import load_workbook

workbook = load_workbook(filename=lab_excel_file)
main_student_hash = {}

if "combined" in workbook.sheetnames:
    workbook.remove(workbook["combined"])

for wb_tab in workbook.sheetnames:
    select_active_tab(workbook, wb_tab)
    active_sheet = workbook.active
    
    header_row = active_sheet[1]
    active_sheet_headers = [cell.value for cell in header_row]

    calculate_number_of_labs_attended(active_sheet, wb_tab)


if "combined" not in active_sheet_headers:
    workbook.create_sheet("combined", 0)
    
select_active_tab(workbook, "combined")
workbook.active.cell(row=1, column=1).value = "Student"
workbook.active.cell(row=1, column=2).value = "ID"
workbook.active.cell(row=1, column=3).value = "SIS User ID"
workbook.active.cell(row=1, column=4).value = "SIS Login ID"
workbook.active.cell(row=1, column=5).value = "Root Account"
workbook.active.cell(row=1, column=6).value = "Section"


all_lab_sheetnames = workbook.sheetnames
all_lab_sheetnames.remove('combined')
all_lab_dates = []

for index, lab in enumerate(all_lab_sheetnames):
    all_lab_dates.append(lab)
    if lab in key_lab_dates:
        all_lab_dates.append('Points earned by ' + lab)
    

student_index = 2

for key, value in main_student_hash.items():
    total_semester_lab_grade = 20
    num_of_missing_labs = 0
    
    workbook.active.cell(row=student_index, column=1).value = value['student']
    workbook.active.cell(row=student_index, column=2).value = key
    workbook.active.cell(row=student_index, column=3).value = value['sis_user_id']
    workbook.active.cell(row=student_index, column=4).value = value['sis_login_id']
    workbook.active.cell(row=student_index, column=5).value = value['root_account']
    workbook.active.cell(row=student_index, column=6).value = value['section']
    
    
    for index, lab in enumerate(all_lab_dates):
        workbook.active.cell(row=1, column=7 + index).value = lab + ' labs'

        
        if 'Points earned by' not in lab and lab not in value:
            workbook.active.cell(row=student_index, column=7 + index).value = -100
            continue


        if 'Points earned by' in lab:
            lab_name = lab.replace('Points earned by ', '')
            if lab_name not in value:
                workbook.active.cell(row=student_index, column=7 + index).value = 'N/A'
                continue
            workbook.active.cell(row=student_index, column=7 + index).value = lab_points_earned_by_deadline(value[lab_name][lab_name + ' # of labs attended'],
                                                                                                                     (key_lab_dates.index(lab_name) + 1))
            
        if lab in key_lab_dates:
            total_semester_lab_grade = total_semester_lab_grade - lab_attendance_deduction_after_key_lab(value[lab][lab + ' # of labs attended'],
                                                                                                         (key_lab_dates.index(lab) + 1))
        
        if 'Points earned by' not in lab:        
            workbook.active.cell(row=student_index, column=7 + index).value = total_semester_lab_grade
        
    student_index = student_index + 1

from datetime import datetime

new_file_created_name = datetime.today().strftime('%Y-%m-%d') + "_" + lab_excel_file

workbook.save(filename=new_file_created_name)
print("Excel file {} updated and saved".format(new_file_created_name))
